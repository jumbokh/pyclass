from openpyxl import load_workbook
import pandas as pd
import numpy as np

class CapitalBudget:
    
    def npv(k,s):  # k為折現率 s為現金流量
        return np.npv(k,s)  # numpy有計算NPV的功能

    def pi(k,s):
        sum=0
        for i in range(1,len(s)):
            sum=sum+(s[i]/pow((1+k),i))
        return sum/abs(s[0])     #abs(s[0])將第0期的現金流量絕對值

    def pvr(k,s):
        P1V1=0
        P2V2=0
        sumP1V1=0
        sumP2V2=0
        for i in range(1,len(s)):
            r1=0.05
            r2=0.055
            F1=1/(pow(1+r1,i-0))
            F2=1/(pow(1+r2,i-0))
            P1V1=s[i]*F1
            P2V2=s[i]*F2
            sumP1V1+=P1V1
            sumP2V2+=P2V2
        return r1+((sumP1V1-abs(s[0]))/(sumP1V1-sumP2V2))*(r2-r1)    #abs(s[0])將第0期的現金流量絕對值

wb=load_workbook('./H_7_7/H_7_7.xlsx')
ws=wb.active
rows=ws.max_row
columns=ws.max_column
list1=[]
list2=[]
list3=[]
for i in range(1,columns+1):
    for j in range(1,rows+1):
        list1.append(ws.cell(column=i,row=j).value)
    list2.append(CapitalBudget.npv(0.05,list1))
    list2.append(CapitalBudget.pi(0.05,list1))
    list2.append(CapitalBudget.pvr(0.05,list1))
    list3.append(list2)
    list1=[]
    list2=[]

frame = pd.DataFrame(list3,columns=['NPV','PI','PVR'],index=['A','B','C'])
print(frame)