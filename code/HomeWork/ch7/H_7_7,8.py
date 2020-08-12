from openpyxl import load_workbook
import numpy as np
import pandas as pd


def npv(k,s):  # k為折現率 s為現金流量
    print('NPV=',np.npv(k,s))
    return np.npv(k,s)  # numpy有計算NPV的功能

def pi(k,s):
    sum=0
    for i in range(1,len(s)):
        sum=sum+(s[i]/pow((1+k),i))
    print('PI=',sum/abs(s[0]) )
    return sum/abs(s[0])     #abs(s[0])將第0期的現金流量絕對值

def pvr(k,s):
    P1V1=0
    P2V2=0
    sumP1V1=0
    sumP2V2=0
    for i in range(1,len(s)):
        r1=0.05
        r2=0.055
        F1=1/(pow(1+r1,i-0))
        F2=1/(pow(1+r2,i-0))
        P1V1=s[i]*F1
        P2V2=s[i]*F2
        sumP1V1+=P1V1
        sumP2V2+=P2V2
    print('pvr=',r1+((sumP1V1-abs(s[0]))/(sumP1V1-sumP2V2))*(r2-r1))
    return r1+((sumP1V1-abs(s[0]))/(sumP1V1-sumP2V2))*(r2-r1)    #abs(s[0])將第0期的現金流量絕對值


wb=load_workbook('./H_7_7/H_7_7.xlsx')
ws1=wb.worksheets[0]
ws2=wb.worksheets[1]          #把統計結果顯示在sheet2
ws2.title='評估分析'

for row in ws2['A1:C5']:   #可以將sheet2的值從A1:C5都清空
    for cell in row:
        cell.value = None
    
ws2.cell(column=1,row=1).value='方案'
ws2.cell(column=2,row=1).value='NPV'
ws2.cell(column=3,row=1).value='PI'
ws2.cell(column=4,row=1).value='PVR'
ws2.cell(column=1,row=2).value='A'
ws2.cell(column=1,row=3).value='B'
ws2.cell(column=1,row=4).value='C'

rows=ws1.max_row
columns=ws1.max_column

list1=[]
for i in range(1,columns+1):
    for j in range(1,rows+1):
        list1.append(ws1.cell(column=i,row=j).value)
    ws2.cell(column=2,row=i+1).value=npv(0.05,list1)
    ws2.cell(column=3,row=i+1).value=pi(0.05,list1)
    ws2.cell(column=4,row=i+1).value=pvr(0.05,list1)
    if(npv(0.05,list1)>0 & (pi(0.05,list1)>1 or pvr(0.05,list1)>0.05)):
        ws2.cell(column=5,row=i+1).value='值得投資'
    else:
        ws2.cell(column=5,row=i+1).value='不值得投資'
    list1=[]
    
wb.save('./H_7_7/H_7_7.xlsx')