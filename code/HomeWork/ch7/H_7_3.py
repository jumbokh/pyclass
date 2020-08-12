from openpyxl import load_workbook
import numpy as np

wb=load_workbook('./H_7_3/H_7_3.xlsx')
ws1=wb.worksheets[0] 
ws2=wb.worksheets[1]
ws2.title='相關係數' #把工作表2的名稱設為'相關係數'
ws2['A1']='ETFs'

rows=ws1.max_row
cols=ws1.max_column
for i in range(2,cols+1):
    ws2.cell(row=1,column=i).value=ws1.cell(row=1,column=i).value  #將工作表1的公司名稱移到工作表2'相關係數'的列
    ws2.cell(row=i,column=1).value=ws1.cell(row=1,column=i).value  #將工作表1的公司名稱移到工作表2'相關係數'的行

list1=[]
list2=[]
x=2   #把第二攔設為跟其他公司比較的公司 (例:S&P500)
i=2   #把第二欄設為被比較的公司，比較完後會依序增加 (例:S&P500-日本2X-FB上証....)
while i<cols+1:   
    for j in range(2,rows+1):
        list1.append(ws1.cell(column=x,row=j).value)   #x攔為比較的公司，將公司的每行數據存進list1裡
        list2.append(ws1.cell(column=i,row=j).value)   #i攔為被比較的公司，將公司的每行數據存進list2裡
        if(j==rows):                                   #如果每行的數據都成功匯進list1跟list2時
            ws2.cell(column=x,row=i).value=np.corrcoef(list1,list2)[0,1] #corrcoef得到相關係數矩陣，並把值存到ws2的儲存格裡
    list1=[]                #將數據歸零
    list2=[]
    if(x==cols & i==cols):  #如果都成功跑完，程式結束
        break
    elif(i==cols):   #如果比較完，跟其他公司比較的公司換成下一個，並且被比較的公司從頭開始
        x=x+1
        i=2
    else:            #被比較的公司換成下一個
        i=i+1 

wb.save('./H_7_3/H_7_3.xlsx')