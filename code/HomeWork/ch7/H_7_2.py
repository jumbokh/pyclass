from openpyxl import load_workbook
import numpy as np

wb=load_workbook('./H_7_2/H_7_2.xlsx')
ws1=wb.worksheets[0]   #設ws1為工作表1
rows=ws1.max_row     #取最大行
cols=ws1.max_column  #取最大列
print('行=',rows)
print('攔=',cols)
ws1['D1']='相關係數'
list1=[]
list2=[]

for i in range(2,rows+1):
    list1.append(ws1.cell(column=2,row=i).value)  #取第2攔的值進list1裡
    list2.append(ws1.cell(column=3,row=i).value)  #取第3攔的值進list2裡


ws1.cell(row=2,column=4).value=np.corrcoef(list1,list2)[0,1]  #corrcoef得到相關係數矩陣，並把值存到(row=2,column=4)的儲存格裡
print('相關係數=',np.corrcoef(list1,list2)[0,1])


wb.save('./H_7_2/H_7_2.xlsx')