from openpyxl import load_workbook
import numpy as np

wb=load_workbook('./H_7_1/H_7_1.xlsx')
ws1=wb.worksheets[0]
rows=ws1.max_row  #取行的最大長度

ws1.cell(row=1,column=7).value='加權股價指數'
ws1.cell(row=1,column=8).value='工業生產指數'
ws1.cell(row=1,column=9).value='物價'
ws1.cell(row=1,column=10).value='匯率'
ws1.cell(row=1,column=11).value='利率'


for j in range(2,7):  #取2到6欄的迴圈
    list1=[]                 
    for i in range(3,rows+1):  #取3到最大行的迴圈
        list1.append(ws1.cell(row=i,column=j).value) #將資料儲存在list1裡
    for i in range(3,rows+1):
        avg=np.sum(list1)/(rows-2)  #rows-2是因為前2行沒有資料
        # ddof=1 指的是除以N-1的標準差，預設則為ddof=0，是除以N的標準差
        ws1.cell(row=i,column=j+5).value=(ws1.cell(row=i,column=j).value-avg)/np.std(list1, ddof=1) 
    
wb.save('./H_7_1/H_7_1.xlsx')